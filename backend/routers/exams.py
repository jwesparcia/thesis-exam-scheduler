from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from core import get_db
from model import Exam, Subject, Section, Room, Timeslot, Course, YearLevel, Teacher, User, Notification
from room_data import AVAILABLE_EXAM_ROOMS, get_room_names_for_department
from utils.scheduler import generate_exam_schedule
from datetime import datetime
from threading import Lock
from typing import Optional
from .auth import get_current_user, require_role
from utils.logging import log_activity
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/exams", tags=["Exams"])

_generation_progress = {}
_generation_progress_lock = Lock()

# Rate limiter: 1 generation request per user per 60 seconds
_GENERATE_RATE_LIMIT_SECONDS = 60
_generate_last_request: dict = {}  # user_id -> datetime of last accepted request
_generate_rate_limit_lock = Lock()


def _progress_key(current_user: User, job_id: str | None = None):
    return str(job_id or current_user.id)


def _set_generation_progress(key, status, percent, phase, detail=""):
    with _generation_progress_lock:
        _generation_progress[key] = {
            "status": status,
            "percent": max(0, min(100, int(percent))),
            "phase": phase,
            "detail": detail,
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }


def _get_generation_progress(key):
    with _generation_progress_lock:
        return _generation_progress.get(key, {
            "status": "idle",
            "percent": 0,
            "phase": "Idle",
            "detail": "",
            "updated_at": None,
        })


def is_generation_ongoing() -> bool:
    with _generation_progress_lock:
        return any(job.get("status") == "running" for job in _generation_progress.values())


def _format_exam_for_room_status(exam: Exam):
    subject = exam.subject
    section = exam.section
    course = exam.course
    year = exam.year_level
    timeslot = exam.timeslot
    room = exam.room

    if timeslot:
        exam_date = timeslot.date.strftime("%A, %B %d, %Y")
        start_time = timeslot.start_time.strftime("%I:%M %p")
        end_time = timeslot.end_time.strftime("%I:%M %p")
    else:
        exam_date = "-"
        start_time = "-"
        end_time = "-"

    return {
        "id": exam.id,
        "subject_code": subject.code if subject else "-",
        "subject_name": subject.name if subject else "-",
        "section_name": section.name if section else "-",
        "course_name": course.name if course else "-",
        "department": course.category if course else "-",
        "year_level": year.name if year else "-",
        "semester": exam.semester,
        "status": exam.status,
        "exam_date": exam_date,
        "start_time": start_time,
        "end_time": end_time,
        "room_id": exam.room_id,
        "room": room.name if room else None,
    }

@router.get("/")
def get_exams(
    status: str = Query(None, description="Filter by status (draft or posted)"),
    section_name: str = Query(None, description="Filter by section name (e.g., BSIT 3-201)"),
    course_id: int = Query(None, description="Filter by course ID"),
    year_level_id: int = Query(None, description="Filter by year level ID"),
    semester: int = Query(None, description="Filter by semester"),
    proctor_id: int = Query(None, description="Filter by proctor (teacher) ID"),
    term: str = Query(None, description="Filter by term (e.g. Midterm, Final)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Fetch exams with optional filters. Students can filter by their section.
    Returns joined data so students can see full details.
    """
    print(f"[DEBUG] get_exams called filters: status={status}, section={section_name}, course={course_id}, year={year_level_id}, sem={semester}, term={term}")
    
    query = db.query(Exam).options(
        joinedload(Exam.subject),
        joinedload(Exam.section),
        joinedload(Exam.room),
        joinedload(Exam.timeslot),
        joinedload(Exam.course),
        joinedload(Exam.year_level),
        joinedload(Exam.proctor),
    )

    if status:
        query = query.filter(Exam.status == status)
    
    if section_name:
        # Use has() to filter by relationship without creating duplicates
        query = query.filter(Exam.section.has(name=section_name))

    if course_id:
        query = query.filter(Exam.course_id == course_id)
    
    if year_level_id:
        query = query.filter(Exam.year_level_id == year_level_id)
        
    if semester:
        query = query.filter(Exam.semester == semester)

    if proctor_id:
        query = query.filter(Exam.proctor_id == proctor_id)

    if term:
        query = query.filter(Exam.term == term)

    # Join with Timeslot to order by date/time
    query = query.join(Exam.timeslot).order_by(Timeslot.date, Timeslot.start_time)

    exams = query.all()
    print(f"[DEBUG] Found {len(exams)} exams")
    
    if not exams:
        return []

    result = []
    for e in exams:
        subject = e.subject
        section = e.section
        room = e.room
        timeslot = e.timeslot
        course = e.course
        year = e.year_level
        proctor = e.proctor

        # Format date with day name to match scheduler format
        if timeslot:
            day_name = timeslot.date.strftime("%A")
            date_str = timeslot.date.strftime("%B %d, %Y")
            full_date = f"{day_name}, {date_str}"
        else:
            full_date = "-"

        proctor_name = proctor.name if proctor else "Unassigned"

        result.append({
            "id": e.id,
            "subject_code": subject.code if subject else "-",
            "subject_name": subject.name if subject else "-",
            "exam_type": subject.exam_type if subject else "-",
            "category": subject.category if subject else "-",
            "section_name": section.name if section else "-",
            "course_name": course.name if course else "-",
            "year_level": year.name if year else "-",
            "semester": e.semester,
            "term": e.term or "Midterm",
            "status": e.status or "draft",
            "exam_date": full_date,
            "start_time": timeslot.start_time.strftime("%I:%M %p") if timeslot else "-",
            "end_time": timeslot.end_time.strftime("%I:%M %p") if timeslot else "-",
            "room": room.name if room else "-",
            "proctor": proctor_name,
            "proctor_name": proctor_name,
            "proctor_attendance": e.proctor_attendance or "pending",
        })

    return result

@router.get("/subjects")
def get_department_subjects(
    department: str = Query("College", description="Department category (e.g., College or SHS)"),
    semester: int = Query(1, description="Semester (1 or 2)"),
    course_id: Optional[int] = Query(None, description="Optional: filter by specific course ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get all unique written subject names for a specific department and semester.
    Optionally filter by course_id when a specific course is selected.
    """
    query = db.query(Subject.name).join(Course).filter(
        Course.category == department,
        Subject.exam_type == "written",
        Subject.semester == semester
    )
    if course_id:
        query = query.filter(Course.id == course_id)
    subjects = query.distinct().order_by(Subject.name).all()
    
    return [s[0] for s in subjects if s[0]]


@router.get("/rooms")
def list_rooms(
    department: str = Query(None, description="Filter by department (College or SHS)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return all rooms, optionally filtered by department."""
    query = db.query(Room)
    if department:
        query = query.filter(Room.department == department)
    rooms = query.order_by(Room.name).all()
    return [
        {"id": r.id, "name": r.name, "building": r.building, "capacity": r.capacity, "department": r.department}
        for r in rooms
    ]


@router.post("/rooms")
def create_room(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    if is_generation_ongoing():
        raise HTTPException(status_code=400, detail="Cannot add rooms while schedule generation is ongoing")

    name = payload.get("name")
    building = payload.get("building")
    capacity = payload.get("capacity")
    
    if not name or not building or capacity is None:
        raise HTTPException(status_code=400, detail="Name, building, and capacity are required")
        
    building_upper = str(building).strip().upper()
    if building_upper not in ["B", "C"]:
        raise HTTPException(status_code=400, detail="Building must be B or C")
        
    try:
        capacity_int = int(capacity)
        if capacity_int <= 0:
            raise ValueError()
    except ValueError:
        raise HTTPException(status_code=400, detail="Capacity must be a positive integer")
        
    # Infer department from building
    dept = "College" if building_upper == "B" else "SHS"
    
    # Check if name is taken
    existing = db.query(Room).filter(Room.name == name).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Room {name} already exists")
        
    new_room = Room(
        name=name,
        building=building_upper,
        capacity=capacity_int,
        department=dept
    )
    db.add(new_room)
    db.commit()
    db.refresh(new_room)
    
    try:
        log_activity(
            db,
            user_id=current_user.id,
            action="CREATE_ROOM",
            details=f"Created room {new_room.name} (Building: {new_room.building}, Capacity: {new_room.capacity})"
        )
    except Exception as e:
        print(f"Error logging room creation: {e}")

    return {
        "message": "Room created successfully",
        "room": {
            "id": new_room.id,
            "name": new_room.name,
            "building": new_room.building,
            "capacity": new_room.capacity,
            "department": new_room.department
        }
    }


@router.delete("/rooms/{room_id}")
def delete_room(
    room_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    if is_generation_ongoing():
        raise HTTPException(status_code=400, detail="Cannot delete rooms while schedule generation is ongoing")
        
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
        
    # Unassign this room from any existing exams
    db.query(Exam).filter(Exam.room_id == room_id).update({Exam.room_id: None}, synchronize_session=False)
    
    db.delete(room)
    db.commit()
    
    try:
        log_activity(
            db,
            user_id=current_user.id,
            action="DELETE_ROOM",
            details=f"Deleted room {room.name} (ID: {room_id})"
        )
    except Exception as e:
        print(f"Error logging room deletion: {e}")
        
    return {"message": "Room deleted successfully"}



@router.get("/rooms/status")
def get_room_status(
    department: str = Query("College", description="College, SHS, or All"),
    semester: Optional[int] = Query(None, description="Optional semester filter"),
    status: str = Query("all", description="draft, posted, or all"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    selected_department = department if department in ["College", "SHS"] else "All"
    
    # Query rooms from the database directly
    rooms_query = db.query(Room)
    if selected_department != "All":
        rooms_query = rooms_query.filter(Room.department == selected_department)
    rooms = rooms_query.order_by(Room.name).all()
    
    allowed_room_names = {room.name for room in rooms}

    exam_query = db.query(Exam).options(
        joinedload(Exam.subject),
        joinedload(Exam.section),
        joinedload(Exam.course),
        joinedload(Exam.year_level),
        joinedload(Exam.timeslot),
        joinedload(Exam.room),
    ).join(Course, Exam.course_id == Course.id)

    if selected_department != "All":
        exam_query = exam_query.filter(Course.category == selected_department)
    if semester:
        exam_query = exam_query.filter(Exam.semester == semester)
    if status in ["draft", "posted"]:
        exam_query = exam_query.filter(Exam.status == status)

    exams = exam_query.all()

    room_bookings = {room.id: [] for room in rooms}
    unassigned_exams = []
    wrong_building_exams = []
    conflict_map = {}

    for exam in exams:
        if not exam.room_id or not exam.room:
            unassigned_exams.append(exam)
            continue

        if selected_department != "All" and exam.room.name not in allowed_room_names:
            wrong_building_exams.append(exam)

        if exam.room_id in room_bookings:
            room_bookings[exam.room_id].append(exam)

        if exam.timeslot_id:
            conflict_key = (exam.room_id, exam.timeslot_id)
            conflict_map.setdefault(conflict_key, []).append(exam)

    conflict_exam_ids = set()
    conflicts = []
    for conflict_exams in conflict_map.values():
        if len(conflict_exams) <= 1:
            continue
        for exam in conflict_exams:
            conflict_exam_ids.add(exam.id)
        first_exam = conflict_exams[0]
        conflicts.append({
            "room": first_exam.room.name if first_exam.room else "-",
            "exam_date": first_exam.timeslot.date.strftime("%A, %B %d, %Y") if first_exam.timeslot else "-",
            "start_time": first_exam.timeslot.start_time.strftime("%I:%M %p") if first_exam.timeslot else "-",
            "end_time": first_exam.timeslot.end_time.strftime("%I:%M %p") if first_exam.timeslot else "-",
            "exams": [_format_exam_for_room_status(exam) for exam in conflict_exams],
        })

    room_rows = []
    for room in rooms:
        bookings = room_bookings.get(room.id, [])
        has_conflict = any(exam.id in conflict_exam_ids for exam in bookings)
        room_rows.append({
            "id": room.id,
            "name": room.name,
            "building": room.building or room.name[:1],
            "department": room.department or "Unknown",
            "capacity": room.capacity or 40,
            "status": "conflict" if has_conflict else "in_use" if bookings else "available",
            "booking_count": len(bookings),
            "draft_count": sum(1 for exam in bookings if exam.status == "draft"),
            "posted_count": sum(1 for exam in bookings if exam.status == "posted"),
            "bookings": [_format_exam_for_room_status(exam) for exam in sorted(
                bookings,
                key=lambda e: (
                    e.timeslot.date if e.timeslot else datetime.max.date(),
                    e.timeslot.start_time if e.timeslot else datetime.max.time(),
                    e.section.name if e.section else "",
                )
            )],
        })

    return {
        "filters": {
            "department": selected_department,
            "semester": semester,
            "status": status,
        },
        "summary": {
            "total_rooms": len(rooms),
            "in_use_rooms": sum(1 for room in room_rows if room["status"] in ["in_use", "conflict"]),
            "available_rooms": sum(1 for room in room_rows if room["status"] == "available"),
            "conflict_rooms": sum(1 for room in room_rows if room["status"] == "conflict"),
            "total_exams": len(exams),
            "assigned_exams": sum(1 for exam in exams if exam.room_id and exam.room),
            "unassigned_exams": len(unassigned_exams),
            "wrong_building_exams": len(wrong_building_exams),
            "conflicts": len(conflicts),
        },
        "rooms": room_rows,
        "unassigned_exams": [_format_exam_for_room_status(exam) for exam in unassigned_exams],
        "wrong_building_exams": [_format_exam_for_room_status(exam) for exam in wrong_building_exams],
        "conflicts": conflicts,
    }


@router.get("/generate/progress")
def get_generate_progress(
    job_id: str = Query(None, description="Generation job id returned or sent with /exams/generate"),
    current_user: User = Depends(require_role(["admin"]))
):
    return _get_generation_progress(_progress_key(current_user, job_id))

@router.post("/generate/cancel")
def cancel_schedule_generation(
    job_id: Optional[str] = Query(None, description="Generation job id to cancel"),
    current_user: User = Depends(require_role(["admin"]))
):
    key = _progress_key(current_user, job_id)
    with _generation_progress_lock:
        current = _generation_progress.get(key)
        if current and current.get("status") == "running":
            _generation_progress[key]["status"] = "cancelled"
            _generation_progress[key]["phase"] = "Cancelling..."
            _generation_progress[key]["detail"] = "Cancellation requested by user."
            return {"message": "Generation cancellation requested."}
        else:
            # Try to cancel any running job
            cancelled_any = False
            for k, v in _generation_progress.items():
                if v.get("status") == "running":
                    v["status"] = "cancelled"
                    v["phase"] = "Cancelling..."
                    v["detail"] = "Cancellation requested by user."
                    cancelled_any = True
            if cancelled_any:
                return {"message": "Generation cancellation requested."}
            return {"message": "No active schedule generation job to cancel."}

@router.post("/generate")
def generate_schedule(
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    """
    Trigger the automatic scheduling for ALL courses based on distribution rules.
    Optionally accepts start_date (YYYY-MM-DD), end_date, department, semester, and excluded_subjects in the body.
    """
    # --- Rate Limiting: 1 request per 60 seconds per user ---
    now = datetime.utcnow()
    with _generate_rate_limit_lock:
        last = _generate_last_request.get(current_user.id)
        if last is not None:
            elapsed = (now - last).total_seconds()
            remaining = _GENERATE_RATE_LIMIT_SECONDS - elapsed
            if remaining > 0:
                raise HTTPException(
                    status_code=429,
                    detail=f"Too many requests. Please wait {int(remaining) + 1} more second(s) before generating again."
                )
        _generate_last_request[current_user.id] = now

    try:
        payload_data = payload if payload else {}
        job_id = _progress_key(current_user, payload_data.get("job_id"))
        _set_generation_progress(job_id, "running", 1, "Preparing schedule", "Starting schedule generation")

        start_date = None
        end_date = None
        
        department = payload_data.get("department", "College")
        semester = payload_data.get("semester", 1)
        term = payload_data.get("term", "Midterm")
        excluded_subjects = payload_data.get("excluded_subjects", [])
        
        start_date_str = payload_data.get("start_date")
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date_str = payload_data.get("end_date")
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        
        # --- Date Validation ---
        if start_date and end_date and end_date < start_date:
            _set_generation_progress(job_id, "failed", 0, "Generation failed", "End date must be after or equal to start date.")
            raise HTTPException(status_code=400, detail="End date cannot be before start date.")

        # --- Guard: Ensure at least one student account exists ---
        student_count = db.query(User).filter(User.role == "student").count()
        if student_count == 0:
            _set_generation_progress(job_id, "failed", 0, "Generation blocked", "No student accounts found.")
            raise HTTPException(
                status_code=400,
                detail="Cannot generate schedule: no student accounts have been uploaded. Please import students first."
            )

        def _on_progress(progress):
            current = _get_generation_progress(job_id)
            if current.get("status") == "cancelled":
                raise RuntimeError("Schedule generation cancelled by user.")
            _set_generation_progress(
                job_id,
                "running",
                progress.get("percent", 0),
                progress.get("phase", "Generating schedule"),
                progress.get("detail", "")
            )

        result = generate_exam_schedule(
            db, 
            start_date=start_date, 
            end_date=end_date, 
            department=department, 
            semester=semester,
            excluded_subjects=excluded_subjects,
            term=term,
            progress_callback=_on_progress
        )
        
        total = result["total_exams"]
        assigned = result["assigned_proctors"]
        unassigned = result["unassigned"]
        unassigned_rooms = result.get("unassigned_rooms", 0)
        
        message = f"Schedule generated! {total} exams created for {department}. "
        if assigned > 0:
            message += f"{assigned} exams assigned a proctor. "
        if unassigned > 0:
            message += f"{unassigned} exams have no proctor due to insufficient availability or limits."
        else:
            message += "All exams have a proctor assigned."
        if unassigned_rooms > 0:
            message += f" {unassigned_rooms} exams still need rooms."
        
        log_activity(db, current_user.id, "EXAM_GENERATE", f"Dept: {department}, Sem: {semester}", None)
        _set_generation_progress(job_id, "completed", 100, "Schedule generated", message)
        return {"message": message, "job_id": job_id}
    except Exception as e:
        db.rollback()
        if "job_id" in locals():
            current_progress = _get_generation_progress(job_id)
            if current_progress.get("status") == "cancelled" or "cancelled" in str(e).lower():
                _set_generation_progress(
                    job_id,
                    "cancelled",
                    0,
                    "Generation cancelled",
                    "Schedule generation was cancelled by the user."
                )
                raise HTTPException(status_code=400, detail="Schedule generation was cancelled by the user.")
            else:
                _set_generation_progress(
                    job_id,
                    "failed",
                    current_progress.get("percent", 0),
                    "Generation failed",
                    str(e),
                )
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))
        
@router.post("/post")
def post_exams(
    course_id: Optional[int] = Query(None, description="Course ID to post exams for"),
    year_level_id: Optional[int] = Query(None, description="Year level ID to post exams for"),
    semester: int = Query(..., description="Semester to post exams for"),
    department: Optional[str] = Query(None, description="Department (College or SHS)"),
    term: Optional[str] = Query(None, description="Term to post exams for"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    """
    Post draft or saved exams with optional filters (course, year, department, term).
    If no course or year is specified, posts all drafts/saved for the given semester and department.
    """
    if is_generation_ongoing():
        raise HTTPException(status_code=400, detail="Cannot post exams while schedule generation is ongoing")
    query = db.query(Exam).options(
        joinedload(Exam.subject),
        joinedload(Exam.section),
        joinedload(Exam.timeslot)
    ).filter(
        Exam.status.in_(["draft", "saved"]),
        Exam.semester == semester
    )
    
    if course_id:
        query = query.filter(Exam.course_id == course_id)
    if year_level_id:
        query = query.filter(Exam.year_level_id == year_level_id)
    if department:
        query = query.join(Course).filter(Course.category == department)
    if term:
        query = query.filter(Exam.term == term)
        
    latest_drafts = query.all()
    
    if not latest_drafts:
        raise HTTPException(status_code=404, detail="No draft or saved exams found to post")

    for exam in latest_drafts:
        exam.status = "posted"
        if exam.proctor_id:
            proctor_user = db.query(User).filter(User.proctor_id == exam.proctor_id).first()
            if proctor_user:
                sub_code = exam.subject.code if exam.subject else "-"
                sub_name = exam.subject.name if exam.subject else "-"
                sec_name = exam.section.name if exam.section else "-"
                
                date_str = ""
                time_str = ""
                if exam.timeslot:
                    day_name = exam.timeslot.date.strftime("%A")
                    formatted_date = exam.timeslot.date.strftime("%B %d, %Y")
                    date_str = f" on {day_name}, {formatted_date}"
                    time_str = f" at {exam.timeslot.start_time.strftime('%I:%M %p')} - {exam.timeslot.end_time.strftime('%I:%M %p')}"
                
                msg = f"You have been assigned to proctor {sub_name} ({sub_code}) for section {sec_name}{date_str}{time_str}."
                notification = Notification(
                    user_id=proctor_user.id,
                    message=msg,
                    type="info",
                    related_id=exam.id
                )
                db.add(notification)
                
    db.commit()

    log_activity(db, current_user.id, "EXAM_POST", f"Course: {course_id}, Year: {year_level_id}, Sem: {semester}, Dept: {department}")
    return {"message": f"✅ Successfully posted {len(latest_drafts)} exams."}

@router.post("/save")
def save_exams(
    course_id: Optional[int] = Query(None, description="Course ID to save exams for"),
    year_level_id: Optional[int] = Query(None, description="Year level ID to save exams for"),
    semester: int = Query(..., description="Semester to save exams for"),
    department: Optional[str] = Query(None, description="Department (College or SHS)"),
    term: Optional[str] = Query(None, description="Term to save exams for"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    """
    Save draft exams by updating status to 'saved'.
    If no course or year is specified, saves all drafts for the given semester, department, and term.
    """
    if is_generation_ongoing():
        raise HTTPException(status_code=400, detail="Cannot save exams while schedule generation is ongoing")
    query = db.query(Exam).filter(
        Exam.status == "draft",
        Exam.semester == semester
    )
    
    if course_id:
        query = query.filter(Exam.course_id == course_id)
    if year_level_id:
        query = query.filter(Exam.year_level_id == year_level_id)
    if department:
        query = query.join(Course).filter(Course.category == department)
    if term:
        query = query.filter(Exam.term == term)
        
    latest_drafts = query.all()
    
    if not latest_drafts:
        raise HTTPException(status_code=404, detail="No draft exams found to save")

    for exam in latest_drafts:
        exam.status = "saved"
                
    db.commit()

    log_activity(db, current_user.id, "EXAM_SAVE", f"Course: {course_id}, Year: {year_level_id}, Sem: {semester}, Dept: {department}, Term: {term}")
    return {"message": f"✅ Successfully saved {len(latest_drafts)} exams."}

@router.get("/download")
def download_exam_schedule(
    status: str = Query(None, description="Filter by status (draft or posted)"),
    course_id: int = Query(None, description="Filter by course ID"),
    year_level_id: int = Query(None, description="Filter by year level ID"),
    semester: int = Query(None, description="Filter by semester"),
    department: str = Query(None, description="Filter by department (College or SHS)"),
    term: Optional[str] = Query(None, description="Filter by term (e.g. Midterm, Final)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    """
    Download master exam schedules as a unified Excel (.xlsx) file.
    All exams are arranged in a single sheet, sorted chronologically.
    """
    query = db.query(Exam).options(
        joinedload(Exam.subject),
        joinedload(Exam.section),
        joinedload(Exam.room),
        joinedload(Exam.timeslot),
        joinedload(Exam.course),
        joinedload(Exam.year_level),
        joinedload(Exam.proctor),
    )

    if status:
        query = query.filter(Exam.status == status)
    if course_id:
        query = query.filter(Exam.course_id == course_id)
    if year_level_id:
        query = query.filter(Exam.year_level_id == year_level_id)
    if semester:
        query = query.filter(Exam.semester == semester)
    if department:
        query = query.join(Course, Exam.course_id == Course.id).filter(Course.category == department)
    if term:
        query = query.filter(Exam.term == term)

    query = query.join(Timeslot, Exam.timeslot_id == Timeslot.id)
    exams = query.all()

    # Sort exams chronologically: Date -> Start Time -> Course Name -> Section Name
    def sort_key(e):
        t_date = e.timeslot.date if e.timeslot else datetime.max.date()
        t_start = e.timeslot.start_time if e.timeslot else datetime.max.time()
        c_name = e.course.name if e.course else ""
        y_name = e.year_level.name if e.year_level else ""
        s_name = e.section.name if e.section else ""
        return (t_date, t_start, c_name, y_name, s_name)

    exams.sort(key=sort_key)

    # --- Build Excel workbook ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Colour palette
    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark navy
    ACCENT_FILL  = PatternFill("solid", fgColor="EBF3FB")   # light blue row
    TITLE_FILL   = PatternFill("solid", fgColor="2563EB")   # bright blue
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    CELL_FONT    = Font(name="Calibri", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="B0C4DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COL_HEADERS = ["#", "Date", "Time", "Course", "Year Level", "Section", "Term", "Subject Code", "Subject Name", "Room", "Proctor", "Status"]
    COL_WIDTHS  = [5,   22,     20,     12,       14,           14,        12,     14,             32,             12,     28,        10]

    # Create sheet
    dept_name = department if department else "All"
    ws = wb.create_sheet(title="Master Exam Schedule")

    # ---- Title row ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COL_HEADERS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Master Exam Schedule — {dept_name} Department"
    if semester:
        title_cell.value += f"  |  Semester {semester}"
    if term:
        title_cell.value += f"  |  {term}"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    title_cell.border = border
    ws.row_dimensions[1].height = 28

    # ---- Column headers ----
    for col_idx, header in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = border
    ws.row_dimensions[2].height = 22

    if not exams:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COL_HEADERS))
        cell = ws.cell(row=3, column=1, value="No exam schedules found for the selected filters.")
        cell.font = Font(name="Calibri", bold=True, color="FF0000", size=11)
        cell.alignment = CENTER
        cell.border = border
        ws.row_dimensions[3].height = 20
    else:
        # ---- Data rows ----
        for row_num, exam in enumerate(exams, start=1):
            ws_row = row_num + 2  # offset for title + header rows
            fill = ACCENT_FILL if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            timeslot = exam.timeslot
            date_str = timeslot.date.strftime("%A, %B %d, %Y") if timeslot else "-"
            time_str = (
                f"{timeslot.start_time.strftime('%I:%M %p')} – {timeslot.end_time.strftime('%I:%M %p')}"
                if timeslot else "-"
            )
            course_name = exam.course.name if exam.course else "-"
            year_level_name = exam.year_level.name if exam.year_level else "-"
            section_name = exam.section.name if exam.section else "-"
            subject_code = exam.subject.code if exam.subject else "-"
            subject_name = exam.subject.name if exam.subject else "-"
            room_name = exam.room.name if exam.room else "No Room"
            proctor_name = exam.proctor.name if exam.proctor else "Unassigned"

            row_values = [
                row_num,
                date_str,
                time_str,
                course_name,
                year_level_name,
                section_name,
                (exam.term or "Midterm").capitalize(),
                subject_code,
                subject_name,
                room_name,
                proctor_name,
                (exam.status or "-").capitalize(),
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=ws_row, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.fill = fill
                cell.alignment = CENTER if col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 10, 12] else LEFT
                cell.border = border

            ws.row_dimensions[ws_row].height = 18

    # ---- Column widths ----
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes below headers
    ws.freeze_panes = ws.cell(row=3, column=1)

    # Stream the file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    dept_label = f"_{department}" if department else ""
    sem_label = f"_Sem{semester}" if semester else ""
    filename = f"MasterExamSchedule{dept_label}{sem_label}.xlsx"

    log_activity(db, current_user.id, "EXAM_DOWNLOAD", f"Downloaded schedule: dept={department}, sem={semester}")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/clear")
def clear_exams(
    department: Optional[str] = Query(None, description="Department to clear (College/SHS)"),
    semester: Optional[int] = Query(None, description="Semester to clear (1/2)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin"]))
):
    """
    Delete exams. Optionally filters by department and semester.
    """
    if is_generation_ongoing():
        raise HTTPException(status_code=400, detail="Cannot delete exams while schedule generation is ongoing")
        
    query = db.query(Exam)
    
    if department:
        query = query.join(Course).filter(Course.category == department)
        
    if semester:
        query = query.filter(Exam.semester == semester)
        
    exams_to_delete = query.all()
    count = len(exams_to_delete)
    for exam in exams_to_delete:
        db.delete(exam)
    db.commit()
    
    log_activity(db, current_user.id, "EXAM_CLEAR", f"Deleted {count} exams (dept={department}, sem={semester})")
    return {"message": f"🧹 Deleted {count} exams."}
