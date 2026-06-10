"""
schedule_translator.py
======================
Converts a proctor's Excel teaching schedule into a student-friendly plain-text format.

Color coding used in the Excel file:
  - Yellow  : Subject block (Lab or Lecture)
  - Orange  : Break (AM break, PM break, Lunch)
  - Cyan/Blue: Admin hours
  - Black (Theme 1): School-wide break
  - Green   : Consultation hours
  - No fill : Empty / free slot (ignored)

Each block's start and end times are determined by grouping consecutive
rows that share the same background color.

Output format:
    Monday
    7:00 AM – 9:00 AM — Computer Programming 2 (Lab) — Room: RMB409
    10:00 AM – 11:00 AM — Break
    ...
"""

import io
from datetime import datetime, time as _time
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Subject abbreviation → full name translation table
# ---------------------------------------------------------------------------
SUBJECT_TRANSLATIONS = {
    "COMPROG 2": "Computer Programming 2",
    "COMPROG": "Computer Programming",
    "WEBSYS": "Web Systems & Technologies",
    "FUND OF WEB": "Fundamentals of Web Programming",
    "INFO MNGT": "Information Management",
    "PROG LANG": "Programming Language",
    "DISCMATH": "Discrete Mathematics",
    "ORGSTUDY": "Organization and Management",
    "SOFTENG": "Software Engineering",
    "DBMS": "Database Management Systems",
    "ITEC": "IT Elective Course",
    "HUMSS": "Humanities and Social Sciences",
    "MATH": "Mathematics",
    "ENG DATA ANALYSIS": "Engineering Data Analysis",
    "ENG": "English",
    "PHYS": "Physics",
    "CHEM": "Chemistry",
    "ETHICS": "Ethics",
    "RIZAL": "Life and Works of Rizal",
    "NSTP": "National Service Training Program",
    "PE": "Physical Education",
    "COMP PROD": "Computer Production",
    "DIGI SIGNAL": "Digital Signal Processing",
    "EMTECH": "Emerging Technologies",
}

# Day ordering and display names
DAY_ORDER = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
DAY_DISPLAY = {
    "MONDAY": "Monday",
    "TUESDAY": "Tuesday",
    "WEDNESDAY": "Wednesday",
    "THURSDAY": "Thursday",
    "FRIDAY": "Friday",
    "SATURDAY": "Saturday",
    "SUNDAY": "Sunday",
}

# Partial Excel indexed color palette
INDEXED_COLORS = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF",
    5: "FFFF00", 6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF",
    10: "FF0000", 11: "00FF00", 12: "0000FF", 13: "FFFF00",
    14: "FF00FF", 15: "00FFFF", 16: "800000", 17: "008000",
    18: "000080", 19: "808000", 20: "800080", 21: "008080",
    22: "C0C0C0", 23: "808080", 27: "FFFF99", 40: "FFFF00",
    43: "FFC000", 64: "FFFFFF", 65: "000000",
}


# ---------------------------------------------------------------------------
# Color extraction & classification
# ---------------------------------------------------------------------------

def _classify_cell(cell) -> str | None:
    """
    Classify a cell's type based on its background fill color/theme:
      'subject'      → Yellow cells  (Lab / Lecture)
      'break'        → Orange cells  (AM/PM/Lunch break)
      'admin'        → Cyan/Blue cells (Admin hours)
      'school_break' → Black cells   (School-wide break)
      'consultation' → Green cells   (Consultation hours)
      None           → No fill / white / unknown
    """
    try:
        fill = cell.fill
        if fill is None or fill.fill_type in (None, 'none'):
            return None

        fg = fill.fgColor
        if fg is None:
            return None

        # Handle theme-based colors
        if fg.type == 'theme':
            # Theme 1 is typically Black (School-wide Break)
            if fg.theme == 1:
                return 'school_break'
            # Other theme colors (e.g. Header blues) are not mapped to schedule elements
            return None

        # Extract RGB hex
        rgb = None
        if fg.type == 'rgb':
            argb = fg.rgb
            if argb and len(argb) == 8:
                rgb = argb[2:]
            elif argb and len(argb) == 6:
                rgb = argb
        elif fg.type == 'indexed':
            rgb = INDEXED_COLORS.get(fg.indexed)

        if not rgb:
            return None

        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)

        # Near-white → treat as empty
        if r > 230 and g > 230 and b > 230:
            return None

        # Black (all channels very low) → school-wide break
        if r < 60 and g < 60 and b < 60:
            return 'school_break'

        # Yellow → subject (R high, G high, B low)
        if r > 220 and g > 220 and b < 80:
            return 'subject'

        # Orange → break (R high, G mid-high, B low)
        # E.g. FFFFC000 (255, 192, 0)
        if r > 220 and 120 <= g <= 210 and b < 80:
            return 'break'

        # Cyan / Sky blue → admin hours
        # E.g. FF00B0F0 (0, 176, 240)
        if r < 100 and g > 130 and b > 200:
            return 'admin'

        # Green → consultation
        if g > 140 and g > r and g > b and b < 120:
            return 'consultation'

        # Any other colored fill — assume subject
        return 'subject'
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Time parsing / formatting helpers
# ---------------------------------------------------------------------------

def _parse_sequential_timeslots(ws, daytime_col, header_row_idx) -> list[tuple[_time | None, _time | None]]:
    """
    Parse the sequential 30-minute timeslots in the 'Day/Time' column.
    Ensures correct AM/PM tracking when times cross the midday boundary.
    """
    timeslots = []
    prev_minutes = 0  # track accumulated minutes from midnight
    
    rows = list(ws.iter_rows())
    data_rows = rows[header_row_idx + 1:]
    
    for ri, row in enumerate(data_rows):
        val = row[daytime_col].value if daytime_col < len(row) else None
        if val is None:
            timeslots.append((None, None))
            continue
        
        s = str(val).strip()
        if '-' not in s:
            timeslots.append((None, None))
            continue
            
        parts = s.split('-', 1)
        
        def to_mins_and_time(raw_str: str) -> tuple[int, _time] | None:
            raw_str = raw_str.strip().upper()
            if not raw_str:
                return None
            
            # Explicit AM/PM
            if 'AM' in raw_str or 'PM' in raw_str:
                dt = datetime.strptime(raw_str, '%I:%M %p')
                t = dt.time()
                return t.hour * 60 + dt.minute, t
                
            try:
                segs = raw_str.replace(':', ' ').split()
                h = int(segs[0])
                m = int(segs[1])
                
                # Check candidates for 12h-to-24h shift
                mins_candidates = []
                if h == 12:
                    mins_candidates.append((720, _time(12, m)))
                    mins_candidates.append((0, _time(0, m)))
                else:
                    mins_candidates.append((h * 60 + m, _time(h, m)))
                    mins_candidates.append(((h + 12) * 60 + m, _time((h + 12) % 24, m)))
                
                # We expect the schedule to move forward sequentially.
                # Allow a small buffer of 15 minutes for overlapping borders.
                valid = [c for c in mins_candidates if c[0] >= prev_minutes - 15]
                if valid:
                    chosen = min(valid, key=lambda x: x[0])
                else:
                    chosen = max(mins_candidates, key=lambda x: x[0])
                    
                return chosen
            except Exception:
                return None

        p0 = to_mins_and_time(parts[0])
        if p0 is None:
            timeslots.append((None, None))
            continue
        prev_minutes = p0[0]
        
        p1 = to_mins_and_time(parts[1])
        if p1 is None:
            timeslots.append((None, None))
            continue
        prev_minutes = p1[0]
        
        timeslots.append((p0[1], p1[1]))
        
    return timeslots


def _fmt(t: _time | None) -> str:
    """Format a time object as '7:00 AM' or '12:30 PM'."""
    if t is None:
        return ''
    h, m = t.hour, t.minute
    ampm = 'AM' if h < 12 else 'PM'
    h12 = h % 12 or 12
    return f'{h12}:{m:02d} {ampm}'


def _translate_subject(raw: str) -> tuple[str, str | None]:
    """
    Try to expand an abbreviation to its full subject name.
    Returns (full_name, matched_abbreviation_or_None).
    """
    upper = raw.upper().strip()
    
    # Sort keys by length descending to match longest abbreviations first
    sorted_abbrevs = sorted(SUBJECT_TRANSLATIONS.keys(), key=len, reverse=True)
    
    for abbrev in sorted_abbrevs:
        full = SUBJECT_TRANSLATIONS[abbrev]
        if upper == abbrev:
            return full, abbrev
        if upper.startswith(abbrev + " "):
            rest = upper[len(abbrev):].strip()
            # Title case the remaining part, but keep Roman numerals / acronyms upper
            rest_parts = [p if p in ("I", "II", "III", "IV", "V") else p.capitalize() for p in rest.split()]
            rest_str = " ".join(rest_parts)
            return f"{full} {rest_str}".strip(), abbrev
            
    return raw, None


# ---------------------------------------------------------------------------
# Main translation function (color-based, grid format)
# ---------------------------------------------------------------------------

def _is_new_subject_start(val: str) -> bool:
    """
    Heuristically detect if a non-empty string in a subject cell starts a new subject block
    rather than belonging to a details row (like section, class type, or room).
    """
    import re
    val_upper = val.upper().strip()
    if not val_upper:
        return False
        
    # 1. Contains a dash separating digits/letters (e.g. "1-202", "3-201", "2-303")
    if re.search(r'\d-\d', val_upper):
        return False
        
    # 2. Known section prefixes (e.g. BMMA, BSCPE, BACOMM, STEM, etc.)
    SECTION_PREFIXES = (
        'BS', 'BMMA', 'BSCPE', 'BACOMM', 'STEM', 'ABM', 'MAWD', 'BSTM',
        'ACT', 'BSEMC', 'G11', 'G12', 'GRADE', 'SHS', 'TVL', 'GAS', 'HUMSS'
    )
    if val_upper.startswith(SECTION_PREFIXES):
        return False
        
    # 3. Type labels
    if val_upper in ('LAB', 'LEC', 'LECTURE', 'LABORATORY'):
        return False
        
    # 4. Room labels
    if val_upper.startswith(('RM', 'RMB', 'RMC', 'LAB', 'LEC')):
        return False
        
    # 5. Is a pure numeric code
    try:
        float(val_upper)
        return False
    except ValueError:
        pass
        
    return True


def translate_grid_schedule_from_bytes(
    content: bytes,
    sheet_name: str | None = None,
    proctor_name: str | None = None,
) -> str:
    """
    Parse a grid-format Excel schedule using cell background colors to identify
    block types and their exact time boundaries.

    Grid format expected:
      - Row 1 (header): 'Day/Time' | 'MONDAY' | 'TUESDAY' | ...
      - Subsequent rows: time slots in 30-min increments ('07:00 - 07:30', ...)
      - Day columns contain subject info split over multiple rows per block

    Returns a plain-text string (one block per line).
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        return f'Error loading workbook: {exc}'

    # ── Select target worksheet ────────────────────────────────────────────
    ws = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        if proctor_name:
            last = proctor_name.strip().split()[-1].lower()
            for sn in wb.sheetnames:
                if sn.upper() in ('BLANK', 'CHANGES', 'SIMS SYNC'):
                    continue
                if last in sn.lower() or proctor_name.lower() in sn.lower():
                    ws = wb[sn]
                    break
        if ws is None:
            for sn in wb.sheetnames:
                if sn.upper() not in ('BLANK', 'CHANGES', 'SIMS SYNC'):
                    ws = wb[sn]
                    break
        if ws is None:
            ws = wb.active

    all_rows = list(ws.iter_rows())
    if not all_rows:
        return ''

    # ── Locate header row (contains 'Day/Time' or 'TIME') ─────────────────
    header_row_idx = None
    daytime_col = None
    day_columns: dict[int, str] = {}  # col_idx → DAY_NAME_UPPER

    for ri, row in enumerate(all_rows):
        for ci, cell in enumerate(row):
            val = str(cell.value).strip().upper() if cell.value is not None else ''
            if 'DAY' in val and 'TIME' in val:
                daytime_col = ci
                header_row_idx = ri
                break
        if daytime_col is not None:
            break

    if daytime_col is None:
        return ''  # Not a grid-format schedule

    # Identify day columns in the header row
    header_row = all_rows[header_row_idx]
    for ci, cell in enumerate(header_row):
        val = str(cell.value).strip().upper() if cell.value is not None else ''
        if val in DAY_ORDER:
            day_columns[ci] = val

    if not day_columns:
        return ''

    # ── Parse time slots sequentially from the Day/Time column ─────────────
    timeslots = _parse_sequential_timeslots(ws, daytime_col, header_row_idx)

    output_lines: list[str] = []
    used_translations: dict[str, str] = {}

    # ── Process each day column in left-to-right / day order ──────────────
    sorted_cols = sorted(day_columns.keys(),
                         key=lambda c: DAY_ORDER.index(day_columns[c]))

    data_rows = all_rows[header_row_idx + 1:]

    for col_idx in sorted_cols:
        day_upper = day_columns[col_idx]
        day_display = DAY_DISPLAY.get(day_upper, day_upper.capitalize())

        # Pair each data cell with its time slot
        day_data = []
        for ri, row in enumerate(data_rows):
            cell = row[col_idx] if col_idx < len(row) else None
            ts = timeslots[ri] if ri < len(timeslots) else (None, None)
            day_data.append((cell, ts))

        # ── Group consecutive same-color rows into blocks ─────────────────
        blocks = []
        cur_type = None
        cur_block = None

        for cell, ts in day_data:
            cell_type = _classify_cell(cell)
            val = str(cell.value).strip() if (cell is not None and cell.value is not None) else ''

            is_new_subject = (cell_type == 'subject' and _is_new_subject_start(val))

            if cell_type is not None and cell_type == cur_type and not is_new_subject:
                cur_block['values'].append(val)
                cur_block['slots'].append(ts)
            else:
                if cur_block is not None and cur_type is not None:
                    blocks.append(cur_block)
                if cell_type is not None:
                    cur_block = {'type': cell_type, 'values': [val], 'slots': [ts]}
                    cur_type = cell_type
                else:
                    cur_block = None
                    cur_type = None

        if cur_block is not None and cur_type is not None:
            blocks.append(cur_block)

        # ── Convert blocks to output lines ────────────────────────────────
        day_lines = []

        for block in blocks:
            btype = block['type']
            slots = block['slots']
            values = block['values']

            # Compute overall time range for this block
            start_t = next((s for s, _ in slots if s is not None), None)
            end_t   = next((e for _, e in reversed(slots) if e is not None), None)
            if start_t is None or end_t is None:
                continue

            time_str = f'{_fmt(start_t)} – {_fmt(end_t)}'
            non_empty = [v for v in values if v and v.lower() not in ('nan', '')]

            if btype == 'subject':
                subject_raw = non_empty[0] if non_empty else '(Unknown Subject)'
                section     = non_empty[1] if len(non_empty) > 1 else ''
                class_type  = ''
                room        = ''

                if len(non_empty) >= 3:
                    candidate = non_empty[2].upper()
                    if candidate in ('LAB', 'LEC', 'LECTURE', 'LABORATORY'):
                        class_type = 'Lab' if candidate in ('LAB', 'LABORATORY') else 'Lecture'
                        room = non_empty[3] if len(non_empty) > 3 else ''
                    else:
                        # No type row — third non-empty value is the room
                        room = non_empty[2]

                full_subject, abbrev = _translate_subject(subject_raw)
                if abbrev:
                    used_translations[abbrev] = SUBJECT_TRANSLATIONS[abbrev]

                type_str = f' ({class_type})' if class_type else ''
                room_str = f' — Room: {room}' if room else ''
                day_lines.append(f'{time_str} — {full_subject}{type_str}{room_str}')

            elif btype == 'break':
                joined = ' '.join(values).upper()
                if 'LUNCH' in joined:
                    label = 'Lunch Break'
                elif 'AM' in joined:
                    label = 'AM Break'
                elif 'PM' in joined:
                    label = 'PM Break'
                else:
                    label = 'Break'
                day_lines.append(f'{time_str} — {label}')

            elif btype == 'admin':
                day_lines.append(f'{time_str} — Admin Hours')

            elif btype == 'school_break':
                day_lines.append(f'{time_str} — School-wide Break')

            elif btype == 'consultation':
                day_lines.append(f'{time_str} — Consultation Hours')

        if day_lines:
            output_lines.append(day_display)
            output_lines.extend(day_lines)

    # ── Subject translation reference ─────────────────────────────────────
    if used_translations:
        output_lines.append('Subject Translation')
        for abbrev, full in sorted(used_translations.items()):
            output_lines.append(f'- {abbrev}: {full}')

    return '\n'.join(output_lines)


# ---------------------------------------------------------------------------
# Row-based schedule translation (unchanged)
# ---------------------------------------------------------------------------

def translate_row_schedule(df) -> str:
    """
    Translate a row-based Excel schedule (columns: Day, Start Time, End Time,
    Subject, Room) into the same plain-text format.
    """
    import pandas as pd

    day_map = {
        'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3,
        'Friday': 4, 'Saturday': 5, 'Sunday': 6,
    }

    df_clean = df.copy()
    df_clean['_day_idx'] = df_clean['Day'].map(day_map)
    df_clean = df_clean.dropna(subset=['_day_idx'])

    parsed_rows = []
    used_translations = {}

    for idx, row in df_clean.iterrows():
        day_str = str(row['Day']).strip()
        start_val = row['Start Time']
        end_val = row['End Time']
        subject = str(row.get('Subject', '')).strip()
        room = str(row.get('Room', '')).strip()

        def _coerce_time(v):
            if isinstance(v, _time):
                return v
            if isinstance(v, datetime):
                return v.time()
            if isinstance(v, str):
                try:
                    return datetime.strptime(v.strip().upper(), '%I:%M %p').time()
                except Exception:
                    pass
            return None

        start_time = _coerce_time(start_val)
        end_time = _coerce_time(end_val)
        if not start_time or not end_time:
            continue

        full_subject, abbrev = _translate_subject(subject)
        if abbrev:
            used_translations[abbrev] = SUBJECT_TRANSLATIONS[abbrev]

        parsed_rows.append({
            'day': day_str,
            'day_idx': int(row['_day_idx']),
            'start': start_time,
            'end': end_time,
            'subject': full_subject,
            'room': room,
        })

    parsed_rows.sort(key=lambda x: (x['day_idx'], x['start']))

    output = []
    current_day = None
    for r in parsed_rows:
        if r['day'] != current_day:
            current_day = r['day']
            output.append(current_day)
        room_str = f' — Room: {r["room"]}' if r['room'] and r['room'].lower() != 'nan' else ''
        output.append(f'{_fmt(r["start"])} – {_fmt(r["end"])} — {r["subject"]}{room_str}')

    if used_translations:
        output.append('Subject Translation')
        for abbrev, full in sorted(used_translations.items()):
            output.append(f'- {abbrev}: {full}')

    return '\n'.join(output)


# ---------------------------------------------------------------------------
# Legacy wrappers (kept for backward compatibility)
# ---------------------------------------------------------------------------

def translate_grid_schedule(df) -> str:
    """
    Legacy wrapper — kept for backward compatibility.
    Converts a pandas DataFrame back to bytes and uses the color-based translator.
    NOTE: This loses color information! Use translate_grid_schedule_from_bytes() instead.
    """
    return ''
